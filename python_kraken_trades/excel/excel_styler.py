from python_kraken_trades.constants import HEADER_FILL_COLOR
from python_kraken_trades.enums import TradeColumn
from pathlib import Path

def style_excel(output: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    green_fill = PatternFill(start_color=HEADER_FILL_COLOR["positive"], fill_type="solid")
    red_fill = PatternFill(start_color=HEADER_FILL_COLOR["negative"], fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            width = max(len(str(cell.value) or "") for cell in col) + 4
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 40)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = center

        # === Portfolio sheet styling ===
        ws = wb["Portfolio"]
        for row in ws.iter_rows(min_row=2):
            row[0].alignment = left
            row[1].alignment = right
            if str(row[0].value) == "Result":
                row[1].font = bold
                row[1].fill = green_fill if "up" in str(row[1].value) else red_fill

        # === Asset ROI sheet styling ===
        ws = wb["Asset ROI"]
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
        pos_rows = [r for r in rows if r[6].value is not None and r[6].value >= 0]
        neg_rows = [r for r in rows if r[6].value is not None and r[6].value < 0]
        ws.delete_rows(2, ws.max_row)

        def insert_section(title: str, fill: PatternFill, row_list: list, start_row: int):
            ws.insert_rows(start_row)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
            title_cell = ws.cell(row=start_row, column=1)
            title_cell.value = title
            title_cell.font = bold
            title_cell.alignment = center
            title_cell.fill = fill

            for i, row in enumerate(row_list, start=start_row + 1):
                for j, cell in enumerate(row, start=1):
                    new_cell = ws.cell(row=i, column=j, value=cell.value)
                    if j in (7, 8) and isinstance(cell.value, (int, float)):
                        new_cell.fill = green_fill if cell.value >= 0 else red_fill

        if pos_rows:
            insert_section("🟢 Positive ROI Assets 🟢", green_fill, pos_rows, 2)
        if neg_rows:
            insert_section("🔻 Negative ROI Assets 🔻", red_fill, neg_rows, 3 + len(pos_rows))

        # === Individual token sheet styling ===
        left_cols = {
            TradeColumn.UNIQUE_ID.value,
            TradeColumn.DATE.value,
            TradeColumn.PAIR.value,
            TradeColumn.TRADE_TYPE.value,
            TradeColumn.EXECUTION_TYPE.value
        }

        for sheet in wb.sheetnames:
            if sheet in {"Portfolio", "Asset ROI"}:
                continue
            ws = wb[sheet]
            col_names = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2):
                for col_idx, cell in enumerate(row, start=1):
                    col_name = col_names[col_idx - 1] if col_idx - 1 < len(col_names) else ""
                    cell.alignment = left if col_name in left_cols else right

        # === Sheet ordering ===
        wb._sheets.sort(key=lambda s: 0 if s.title == "Portfolio" else (1 if s.title == "Asset ROI" else 2))
        wb.save(output)