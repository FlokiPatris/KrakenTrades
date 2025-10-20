from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from kraken_core import (ExcelStyling, MainSummaryMetrics, TradeMetricsResult,
                         custom_logger)


# =============================================================================
# 🧱 Basic Styling Utilities
# =============================================================================
def _auto_adjust_columns(ws: Worksheet) -> None:
    """Auto-adjust column widths based on header and cell values, with a max width limit."""
    for col in ws.columns:
        try:
            # Include header (first row) + all cell values
            max_width = max(len(str(cell.value) or "") for cell in col)
            col_index = col[0].column
            col_letter: str = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = min(
                max_width + 2, ExcelStyling.MAX_COLUMN_WIDTH
            )
        except Exception as e:
            custom_logger.warning(f"⚠️ Column width adjustment failed: {e}")


def _style_header(ws: Worksheet) -> None:
    """Apply bold + center alignment to header row."""
    for cell in ws[ExcelStyling.HEADER_ROW_INDEX]:
        cell.font = ExcelStyling.BOLD_FONT
        cell.alignment = ExcelStyling.CENTER_ALIGNMENT


def _style_portfolio_sheet(ws: Worksheet) -> None:
    """Apply portfolio-specific formatting."""
    custom_logger.info("📊 Styling Portfolio sheet")
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = ExcelStyling.LEFT_ALIGNMENT
        row[1].alignment = ExcelStyling.RIGHT_ALIGNMENT
        if str(row[0].value) == "Result":
            row[1].font = ExcelStyling.BOLD_FONT
            row[1].fill = (
                ExcelStyling.GREEN_FILL
                if "up" in str(row[1].value).lower()
                else ExcelStyling.RED_FILL
            )


# =============================================================================
# 📈 ROI Sheet Grouping & Conditional Coloring
# =============================================================================
def _insert_roi_section(
    ws: Worksheet,
    title: str,
    fill: PatternFill,
    row_list: List[Iterable[Cell]],
    start_row: int,
) -> None:
    """Insert a titled section with conditional coloring into the Asset ROI sheet."""
    custom_logger.info(f"📌 Inserting section: {title}")
    last_col = ws.max_column
    ws.insert_rows(start_row)
    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=last_col
    )

    title_cell = ws.cell(row=start_row, column=1)
    title_cell.value = title
    title_cell.font = ExcelStyling.BOLD_FONT
    title_cell.alignment = ExcelStyling.CENTER_ALIGNMENT
    title_cell.fill = fill

    for i, row in enumerate(row_list, start=start_row + 1):
        for j, cell in enumerate(row, start=1):
            new_cell = ws.cell(row=i, column=j, value=cell.value)
            if j in (2, 3) and isinstance(cell.value, (int, float)):
                new_cell.fill = (
                    ExcelStyling.GREEN_FILL
                    if cell.value >= 0
                    else ExcelStyling.RED_FILL
                )


def _style_asset_roi_sheet(ws: Worksheet, group_by: str = "roi") -> None:
    """
    Style and group the Asset ROI sheet.

    group_by options:
        - "roi": positive vs negative ROI
        - "remaining_volume": unsold vs sold assets
    """
    custom_logger.info(f"📈 Styling Asset ROI sheet (group_by={group_by})")
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))

    if group_by == "roi":
        try:
            col_idx = headers.index("ROI (%)")
        except ValueError:
            custom_logger.error("❌ ROI (%) column not found")
            return

        pos_rows = [
            r for r in rows if r[col_idx].value is not None and r[col_idx].value >= 0
        ]
        neg_rows = [
            r for r in rows if r[col_idx].value is not None and r[col_idx].value < 0
        ]

        ws.delete_rows(2, ws.max_row)
        if pos_rows:
            _insert_roi_section(
                ws, "🟢 Positive ROI Assets 🟢", ExcelStyling.GREEN_FILL, pos_rows, 2
            )
        if neg_rows:
            _insert_roi_section(
                ws,
                "🔻 Negative ROI Assets 🔻",
                ExcelStyling.RED_FILL,
                neg_rows,
                3 + len(pos_rows),
            )

    elif group_by == "remaining_volume":
        try:
            col_idx = headers.index("Remaining Volume")
        except ValueError:
            custom_logger.error("❌ Remaining Volume column not found")
            return

        unsold_rows = [r for r in rows if r[col_idx].value and r[col_idx].value > 0]
        sold_rows = [
            r for r in rows if r[col_idx].value is not None and r[col_idx].value <= 0
        ]

        ws.delete_rows(2, ws.max_row)
        if unsold_rows:
            _insert_roi_section(
                ws,
                "📦 Unsold Assets (Still Holding)",
                ExcelStyling.GREEN_FILL,
                unsold_rows,
                2,
            )
        if sold_rows:
            _insert_roi_section(
                ws,
                "💰 Sold Assets (Closed Positions)",
                ExcelStyling.RED_FILL,
                sold_rows,
                3 + len(unsold_rows),
            )

    else:
        custom_logger.warning(
            f"⚠️ Unknown group_by value '{group_by}', skipping grouping."
        )


# =============================================================================
# 🧩 Token Sheet Styling
# =============================================================================
def _style_token_sheet(ws: Worksheet) -> None:
    """Apply token-specific formatting."""
    custom_logger.info(f"📄 Styling token sheet: {ws.title}")
    col_names = [cell.value for cell in ws[ExcelStyling.HEADER_ROW_INDEX]]
    for row in ws.iter_rows(min_row=2):
        for col_idx, cell in enumerate(row, start=1):
            col_name = col_names[col_idx - 1] if col_idx - 1 < len(col_names) else ""
            cell.alignment = (
                ExcelStyling.LEFT_ALIGNMENT
                if col_name in ExcelStyling.LEFT_ALIGNED_COLUMNS
                else ExcelStyling.RIGHT_ALIGNMENT
            )


def _reorder_sheets(wb: Workbook) -> None:
    """Reorder sheets: Portfolio → Asset ROI → Tokens."""
    custom_logger.info("🔀 Reordering sheets")
    wb._sheets.sort(
        key=lambda s: (
            0
            if s.title == ExcelStyling.PORTFOLIO_SHEET
            else 1 if s.title == ExcelStyling.ASSET_ROI_SHEET else 2
        )
    )


# =============================================================================
# 🎨 Apply Full Workbook Styling
# =============================================================================
def style_excel(output: Path) -> None:
    """Apply styling to Excel workbook, including Portfolio, ROI, and Token sheets."""
    custom_logger.info(f"📄 Loading workbook: {output}")
    wb = load_workbook(output)

    for ws in wb.worksheets:
        _auto_adjust_columns(ws)
        _style_header(ws)

    # Portfolio
    if ExcelStyling.PORTFOLIO_SHEET in wb.sheetnames:
        _style_portfolio_sheet(wb[ExcelStyling.PORTFOLIO_SHEET])

    # ROI
    if ExcelStyling.ASSET_ROI_SHEET in wb.sheetnames:
        _style_asset_roi_sheet(wb[ExcelStyling.ASSET_ROI_SHEET])

    # Tokens
    for ws in wb.worksheets:
        if ws.title not in {ExcelStyling.PORTFOLIO_SHEET, ExcelStyling.ASSET_ROI_SHEET}:
            _style_token_sheet(ws)

    _reorder_sheets(wb)
    wb.save(output)
    custom_logger.info(f"✅ Workbook saved successfully: {output}")


# =============================================================================
# 🔹 Flatten TradeMetricsResult for ROI Export
# =============================================================================
def flatten_trade_metrics_result(result: TradeMetricsResult) -> dict:
    """
    Flatten TradeMetricsResult (including nested MainSummaryMetrics & MarketData)
    into a single-level dict for pandas DataFrame export.
    Ensures 'token' is the first key.
    """
    flat = {
        "token": result.metrics.token,
        "roi": result.metrics.roi,
        "if_all_sold_now_roi": result.metrics.if_all_sold_now_roi,
    }

    # Then the main summary metrics
    metrics: MainSummaryMetrics = result.metrics
    for key, value in metrics.__dict__.items():
        if key != "market_data" and key != "token":  # avoid duplicate
            flat[key] = value

    # Add nested market data if present
    if metrics.market_data:
        for key, value in metrics.market_data.__dict__.items():
            flat[key] = value

    del flat["price"]  # Remove 'price' to avoid confusion with 'market_price'
    return flat
