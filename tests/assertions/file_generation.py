# =============================================================================
# 📂 Imports
# =============================================================================
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook, Workbook

from kraken_core import custom_logger
from helpers import file_helper


# =============================================================================
# 🛠️ Helper Functions
# =============================================================================
def assert_file_exists(file_path: Path) -> None:
    """
    Assert that a file exists at the given path.

    Raises:
        AssertionError: If the file does not exist.
    """
    assert file_path.exists(), f"❌ Expected file not found: {file_path}"
    custom_logger.info("✅ File exists: %s", file_path)


def assert_valid_excel(file_path: Path) -> None:
    """
    Assert that the Excel file is readable and contains at least one sheet.

    Raises:
        AssertionError: If the file is not a valid Excel workbook or has no sheets.
        FileNotFoundError: If the file does not exist.
    """
    if not file_path.exists():
        raise FileNotFoundError(f"❌ Excel file not found: {file_path}")

    wb: Optional[Workbook] = None
    try:
        wb = load_workbook(file_path)
        assert wb.sheetnames, "❌ Excel file has no sheets"
        custom_logger.info("✅ Excel file is valid and has sheets: %s", wb.sheetnames)
    finally:
        if wb is not None:
            wb.close()


def assert_script_generates_excel(excel_file_name: str) -> None:
    """
    Combined helper: assert the Excel report exists and is valid.

    Args:
        excel_file_name: Name of the Excel file in the uploads directory.

    Raises:
        AssertionError: If file is missing or invalid.
    """
    excel_path: Path = file_helper.uploads_dir / excel_file_name

    assert_file_exists(excel_path)
    assert_valid_excel(excel_path)
