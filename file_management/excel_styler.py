from kraken_core import ExcelStyling, TradeColumn, custom_logger
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def style_excel(output: Path) -> None:
    """
    Apply styling to the Excel workbook at the given path.
    Includes formatting for Portfolio, Asset ROI, and individual token sheets.
    """
    custom_logger.info(f"📄 Loading workbook: {output}")
    wb = load_workbook(output)

    # Define reusable styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    green_fill = PatternFill(start_color=ExcelStyling.HEADER_POSITIVE_FILL, fill_type="solid")
    red_fill = PatternFill(start_color=ExcelStyling.HEADER_NEGATIVE_FILL, fill_type="solid")

    # === General sheet formatting ===
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        custom_logger.info(f"🎨 Formatting sheet: {sheet_name}")

        # Auto-adjust column widths
        for col in ws.columns:
            max_width = max(len(str(cell.value) or "") for cell in col) + 4
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_width, 40)

        # Style header row
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = center

    # === Portfolio sheet styling ===
    if "Portfolio" in wb.sheetnames:
        ws = wb["Portfolio"]
        custom_logger.info("📊 Styling Portfolio sheet")

        for row in ws.iter_rows(min_row=2):
            row[0].alignment = left
            row[1].alignment = right

            if str(row[0].value) == "Result":
                row[1].font = bold
                row[1].fill = green_fill if "up" in str(row[1].value).lower() else red_fill

    # === Asset ROI sheet styling ===
    if "Asset ROI" in wb.sheetnames:
        ws = wb["Asset ROI"]
        custom_logger.info("📈 Styling Asset ROI sheet")

        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
        pos_rows = [r for r in rows if r[6].value is not None and r[6].value >= 0]
        neg_rows = [r for r in rows if r[6].value is not None and r[6].value < 0]

        ws.delete_rows(2, ws.max_row)  # Clear existing data

        def insert_section(title: str, fill: PatternFill, row_list: list, start_row: int):
            custom_logger.info(f"📌 Inserting section: {title}")
            ws.insert_rows(start_row)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)

            title_cell = ws.cell(row=start_row, column=1)
            title_cell.value = title
            title_cell.font = bold
            title_cell.alignment = center
            title_cell.fill = fill

            for i, row in enumerate(row_list, start=start_row + 1):
                for j, cell in enumerate(row, start=1):
                    new_cell = ws.cell(row=i, column=j, value=cell.value)
                    if j in (7, 8) and isinstance(cell.value, (int, float)):
                        new_cell.fill = green_fill if cell.value >= 0 else red_fill

        if pos_rows:
            insert_section("🟢 Positive ROI Assets 🟢", green_fill, pos_rows, 2)
        if neg_rows:
            insert_section("🔻 Negative ROI Assets 🔻", red_fill, neg_rows, 3 + len(pos_rows))

    # === Individual token sheet styling ===
    left_cols = {
        TradeColumn.UNIQUE_ID.value,
        TradeColumn.DATE.value,
        TradeColumn.PAIR.value,
        TradeColumn.TRADE_TYPE.value,
        TradeColumn.EXECUTION_TYPE.value
    }

    for sheet_name in wb.sheetnames:
        if sheet_name in {"Portfolio", "Asset ROI"}:
            continue

        ws = wb[sheet_name]
        custom_logger.info(f"📄 Styling token sheet: {sheet_name}")

        col_names = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for col_idx, cell in enumerate(row, start=1):
                col_name = col_names[col_idx - 1] if col_idx - 1 < len(col_names) else ""
                cell.alignment = left if col_name in left_cols else right

    # === Sheet ordering ===
    custom_logger.info("🔀 Reordering sheets")
    wb._sheets.sort(key=lambda s: 0 if s.title == "Portfolio" else (1 if s.title == "Asset ROI" else 2))

    # Save workbook
    wb.save(output)
    custom_logger.info(f"✅ Workbook saved: {output}")