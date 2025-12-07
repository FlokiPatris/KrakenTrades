# --------------------------------------------------------------------
# 🧰 Environment & Config
# --------------------------------------------------------------------
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import FrozenSet

from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill

from .enums import FolderType, TradeColumn

load_dotenv()  # Load environment variables from .env file


# -------------------------
# Fail-fast helper for env vars
# -------------------------
def get_env_var(name: str, default: str | None = None, required: bool = False) -> str:
    """
    Fetch an environment variable.
    Fail fast if required variable is missing or if no default is provided.
    """
    if name in os.environ:
        return os.environ[name]
    if required:
        raise EnvironmentError(f"Missing required environment variable: {name}")
    if default is not None:
        return default
    raise EnvironmentError(f"No value found for {name} and no default provided")


# --------------------------------------------------------------------
# 🗂 Repository Scanning Configuration
# --------------------------------------------------------------------
@dataclass(frozen=True)
class RepoScanConfig:
    """
    Configuration for scanning repository files.
    """

    SEPARATOR: str = "=" * 80

    # File extensions to include in scan
    INCLUDED_EXTENSIONS: FrozenSet[str] = frozenset(
        [".py", ".txt", ".yml", ".json", ".md", "Dockerfile", "Makefile"]
    )

    # Directories to skip during scanning
    SKIP_DIRS: FrozenSet[str] = frozenset(
        [
            ".git",
            ".hg",
            ".svn",
            "__pycache__",
            ".venv",
            "venv",
            "env",
            "build",
            "dist",
            ".idea",
            ".vscode",
            "tmp",
            "temp",
        ]
    )

    TREE_DEPTH: int = 10
    MAX_FILE_SIZE: int = 10 * 1024 * 1024  # 10 MB
    FALLBACK_CATEGORY: str = "other"

    CATEGORIES_TO_SCAN: FrozenSet[str] = frozenset(
        [".ci", ".github", "scripts", "src", "tests"]
    )
    CATEGORIES_TO_TEST: FrozenSet[str] = frozenset(["e2e", "smoke", "assertions"])


# --------------------------------------------------------------------
# 📁 Paths Configuration
# --------------------------------------------------------------------
@dataclass(frozen=True)
class PathsConfig:
    """
    Centralized configuration for folders and files.
    """

    # --- Files --- #
    KRAKEN_TRADES_PDF: Path = Path(
        get_env_var("KRAKEN_TRADES_PDF", f"{FolderType.DOWNLOADS.value}/trades.pdf")
    )
    PARSED_TRADES_EXCEL: Path = Path(
        get_env_var(
            "PARSED_TRADES_EXCEL",
            f"{FolderType.UPLOADS.value}/kraken_trade_summary.xlsx",
        )
    )

    # --- Folders --- #
    DOWNLOADS_DIR: Path = Path(get_env_var("DOWNLOADS_DIR", FolderType.DOWNLOADS.value))
    UPLOADS_DIR: Path = Path(get_env_var("UPLOADS_DIR", FolderType.UPLOADS.value))
    REPORTS_DIR: Path = Path(
        get_env_var("REPORTS_DIR", f"../{FolderType.REPORTS.value}")
    )

    # --- Repo root --- #
    REPO_ROOT: Path = Path(get_env_var("REPO_ROOT", ".")).resolve()


# --------------------------------------------------------------------
# 🗄 Database / RDS Configuration
# --------------------------------------------------------------------
@dataclass(frozen=True)
class PostgresConfig:
    host: str
    port: int
    dbname: str
    user: str
    password: str = field(repr=False)  # Hide sensitive info in repr

    @staticmethod
    def from_env() -> PostgresConfig:
        """
        Load Postgres configuration from environment variables.
        Fail-fast if required variables are missing.
        """
        host = get_env_var("RDS_HOST", required=True)
        dbname = get_env_var("RDS_DB_NAME", required=True)
        user = get_env_var("RDS_USER", required=True)
        password = get_env_var("RDS_PASSWORD", required=True)
        port = int(get_env_var("RDS_PORT", default="5432"))

        return PostgresConfig(
            host=host, port=port, dbname=dbname, user=user, password=password
        )


# --------------------------------------------------------------------
# 🌐 API Configuration
# --------------------------------------------------------------------
@dataclass(frozen=True)
class KrakenAPI:
    """
    Kraken API configuration.
    Suitable for use in CI/CD and Docker environments.
    """

    URL: str = "https://api.kraken.com/0/public/Ticker"
    TIMEOUT: int = 10  # seconds


# --------------------------------------------------------------------
# 🔍 Regex Patterns
# --------------------------------------------------------------------
class TradeRegex:
    """Precompiled regex patterns for parsing trade data."""

    DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    TRADE = re.compile(
        r"""
        ^(?P<date>\d{4}-\d{2}-\d{2})\s+
        (?P<uid>[A-Z0-9-]+)\s+
        (?P<pair>[A-Z0-9]+/[A-Z0-9]+)\s+
        (?P<type>Buy|Sell)\s+
        (?P<subtype>Limit|Market)\s+
        (?P<price>\d+\.\d+)\s+
        (?P<cost>\d+\.\d+)\s+
        (?P<volume>\d+\.\d+)\s+
        (?P<fee>\d+\.\d+)\s+
        (?P<margin>\d+\.\d+)$
        """,
        re.VERBOSE,
    )
    PAIR_CURRENCY = re.compile(r"/([A-Z]+)")
    PAIR_TOKEN = re.compile(r"^([A-Z0-9]+)/")


# =============================================================================
# 🎨 Styling Configuration
# =============================================================================
@dataclass(frozen=True)
class ExcelStyling:
    # === General constants ===
    MAX_COLUMN_WIDTH: int = 40
    HEADER_ROW_INDEX: int = 1
    PORTFOLIO_SHEET: str = "Portfolio"
    ASSET_ROI_SHEET: str = "Asset ROI"
    SECTION_COLUMNS: int = 13  # Number of columns in merged ROI section header

    # === Header fills (hex strings only) ===
    HEADER_POSITIVE_FILL: str = "C6EFCE"  # Light green
    HEADER_NEGATIVE_FILL: str = "FFC7CE"  # Light red

    # === Fonts & Alignment ===
    BOLD_FONT: Font = Font(bold=True)
    LEFT_ALIGNMENT: Alignment = Alignment(horizontal="left", vertical="center")
    CENTER_ALIGNMENT: Alignment = Alignment(horizontal="center", vertical="center")
    RIGHT_ALIGNMENT: Alignment = Alignment(horizontal="right", vertical="center")

    # === Cell fills ===
    GREEN_FILL: PatternFill = PatternFill(start_color="C6EFCE", fill_type="solid")
    RED_FILL: PatternFill = PatternFill(start_color="FFC7CE", fill_type="solid")

    # === Columns that should be left-aligned ===
    LEFT_ALIGNED_COLUMNS: frozenset[str] = frozenset(
        {
            TradeColumn.UNIQUE_ID.value,
            TradeColumn.DATE.value,
            TradeColumn.PAIR.value,
            TradeColumn.TRADE_TYPE.value,
            TradeColumn.EXECUTION_TYPE.value,
        }
    )


# --------------------------------------------------------------------
# 🔢 Formatting Rules
# --------------------------------------------------------------------
@dataclass(frozen=True)
class FormatRules:
    DECIMAL_PLACES_10: int = 10
    DECIMAL_PLACES_2: int = 2
    DECIMAL_PLACES_8: int = 8


# -----------------------------------------------------------------
# 📊 Main Summary Metric
# -----------------------------------------------------------------
TOKEN_MAP = {
    "ADA/EUR": "cardano",
    "AIOZ/EUR": "aioz-network",
    "AKT/EUR": "akash-network",
    "BEAM/EUR": "beam",
    "BTC/EUR": "bitcoin",
    "DOT/EUR": "polkadot",
    "LTC/EUR": "litecoin",
    "PYTH/EUR": "pyth-network",
    "RENDER/EUR": "render-token",
    "TAO/EUR": "bittensor",
    "WAXL/EUR": "wax",
    "XCN/EUR": "chainx",
    "XRP/EUR": "ripple",
    "ZEUS/EUR": "zeus-network",
    "1000CAT/EUR": "1000cat",
    "1000X/EUR": "1000x-by-virtuals",
    "100YEN/EUR": "100-token",
    "FIGS/EUR": "10-figs",
    "ZHAO/EUR": "-11",
    "GLM/EUR": "golem",
    "DASH/EUR": "dash",
    "ETH/EUR": "ethereum",
    "FET/EUR": "fetch-ai",
    "KAS/EUR": "kaspa",
    "LDO/EUR": "lido-dao",
    "LINK/EUR": "chainlink",
    "MATIC/EUR": "polygon",
    "NANO/EUR": "nano",
    "OCEAN/EUR": "ocean-protocol",
    "SOL/EUR": "solana",
    "XLM/EUR": "stellar",
    "ATOM/EUR": "cosmos",
}
